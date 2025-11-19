import os
import csv
from typing import List, Optional, Protocol

from .files import ensure_dir


class StreamWriter(Protocol):
    """
    公共流式写入接口。

    方法：
    - append_row(row_values): 追加一行并持久化到磁盘
    - close(): 关闭写入会话并释放资源
    """

    def append_row(self, row_values: List[object]) -> None: ...
    def close(self) -> None: ...


class LockFile:
    """
    基于锁文件的跨平台互斥锁，实现并发写入安全。

    实现原理：使用 `os.O_CREAT | os.O_EXCL` 原子方式创建锁文件，成功表示获取锁；
    释放时关闭句柄并删除锁文件。轮询等待锁释放，避免竞争。

    参数：
    - lock_path: 锁文件路径，例如 `result.csv.lock`
    - poll_interval: 轮询间隔秒数，默认 0.1s
    """

    def __init__(self, lock_path: str, poll_interval: float = 0.1):
        import time
        self.lock_path = lock_path
        self.poll_interval = poll_interval
        self._fd: Optional[int] = None
        self._time = time

    def acquire(self) -> None:
        while True:
            try:
                self._fd = os.open(self.lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                os.write(self._fd, f"pid={os.getpid()}\n".encode('utf-8'))
                return
            except FileExistsError:
                self._time.sleep(self.poll_interval)

    def release(self) -> None:
        try:
            if self._fd is not None:
                os.close(self._fd)
                self._fd = None
        finally:
            try:
                if os.path.exists(self.lock_path):
                    os.unlink(self.lock_path)
            except Exception:
                # 尽力而为清理
                pass


class CSVStreamWriter:
    """
    CSV 流式写入器，支持：
    - 原子创建：使用临时文件写入表头并 `os.replace` 原子替换
    - 文件锁：会话期间持有锁，确保并发安全
    - 每行写入后 `flush` + `fsync`，确保持久化

    参数：
    - path: 目标 CSV 文件路径
    - columns: 列名列表；首次创建仅写入表头行

    用法示例：
    >>> writer = CSVStreamWriter('output/foo.csv', ['id', 'input', 'output'])
    >>> writer.append_row([1, 'hello', 'world'])
    >>> writer.close()
    """

    def __init__(self, path: str, columns: List[str]):
        self.path = path
        self.columns = list(columns)
        self.lock = LockFile(path + '.lock')
        self._file = None
        ensure_dir(os.path.dirname(path))
        # 原子写入表头
        tmp = path + '.tmp'
        with open(tmp, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(self.columns)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, path)
        # 会话：持锁并打开追加句柄
        self.lock.acquire()
        self._file = open(self.path, 'a', newline='', encoding='utf-8')
        self._writer = csv.writer(self._file)

    def append_row(self, row_values: List[object]) -> None:
        self._writer.writerow(row_values)
        self._file.flush()
        os.fsync(self._file.fileno())

    def close(self) -> None:
        try:
            try:
                self._file.flush()
                os.fsync(self._file.fileno())
            except Exception:
                pass
            self._file.close()
        finally:
            self.lock.release()


class ExcelStreamWriter:
    """
    Excel(xlsx) 流式写入器，支持：
    - 原子创建：使用 openpyxl 写入表头到临时文件后原子替换
    - 文件锁：会话期间持锁，保证并发安全
    - 每次追加行后保存临时文件并原子替换目标文件

    参数：
    - path: 目标 xlsx 文件路径
    - columns: 列名列表

    用法示例：
    >>> writer = ExcelStreamWriter('output/foo.xlsx', ['id', 'input', 'output'])
    >>> writer.append_row([1, 'hello', 'world'])
    >>> writer.close()
    """

    def __init__(self, path: str, columns: List[str]):
        from openpyxl import Workbook, load_workbook
        self.path = path
        self.columns = list(columns)
        self.lock = LockFile(path + '.lock')
        ensure_dir(os.path.dirname(path))
        wb = Workbook()
        ws = wb.active
        ws.append(self.columns)
        tmp = path + '.tmp'
        wb.save(tmp)
        os.replace(tmp, path)
        # 会话：持锁并加载工作簿
        self.lock.acquire()
        self._wb = load_workbook(self.path)
        self._ws = self._wb.active

    def append_row(self, row_values: List[object]) -> None:
        self._ws.append(row_values)
        tmp = self.path + '.tmp'
        self._wb.save(tmp)
        os.replace(tmp, self.path)

    def close(self) -> None:
        try:
            try:
                tmp = self.path + '.tmp'
                self._wb.save(tmp)
                os.replace(tmp, self.path)
            except Exception:
                pass
            finally:
                try:
                    self._wb.close()
                except Exception:
                    pass
        finally:
            self.lock.release()


def open_stream_writer(file_type: str, path: str, columns: List[str]) -> StreamWriter:
    """
    根据文件类型创建流式写入器，并进行原子表头创建与并发锁定。

    参数：
    - file_type: 文件类型，取值 `csv` 或 `excel`
    - path: 目标文件路径（父目录会自动创建）
    - columns: 列名列表，将作为第一行标题写入

    返回：
    - 实现 `StreamWriter` 协议的对象，支持 `append_row` 与 `close`

    使用示例：
    >>> from utils.streaming import open_stream_writer
    >>> writer = open_stream_writer('csv', 'output/foo.csv', ['id', 'input', 'output'])
    >>> writer.append_row([1, 'hello', 'world'])
    >>> writer.close()
    """
    if file_type == 'csv':
        return CSVStreamWriter(path, columns)
    elif file_type == 'excel':
        return ExcelStreamWriter(path, columns)
    else:
        raise ValueError(f'不支持的文件类型: {file_type}')